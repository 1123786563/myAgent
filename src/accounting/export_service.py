"""
报表导出服务
Report Export Service - Excel, PDF, CSV Export
"""

import io
import csv
from typing import Dict, List, Any, Optional
from datetime import datetime
from decimal import Decimal
from infra.logger import get_logger

log = get_logger("ExportService")


class ReportExportService:
    """报表导出服务"""

    # ==================== CSV 导出 ====================

    def export_to_csv(self, data: Dict[str, Any], report_type: str) -> bytes:
        """导出报表为 CSV 格式"""
        output = io.StringIO()
        writer = csv.writer(output)

        if report_type == "balance-sheet":
            return self._export_balance_sheet_csv(data)
        elif report_type == "income-statement":
            return self._export_income_statement_csv(data)
        elif report_type == "cash-flow":
            return self._export_cash_flow_csv(data)
        elif report_type == "account-balances":
            return self._export_account_balances_csv(data)
        elif report_type == "ledger":
            return self._export_ledger_csv(data)
        else:
            raise ValueError(f"不支持的报表类型: {report_type}")

    def _export_balance_sheet_csv(self, data: Dict) -> bytes:
        """导出资产负债表 CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        # 标题
        writer.writerow([data.get("report_name", "资产负债表")])
        writer.writerow([f"报告期间: {data.get('period', '')}"])
        writer.writerow([f"生成时间: {data.get('generated_at', '')}"])
        writer.writerow([])

        # 资产部分
        writer.writerow(["资产"])
        writer.writerow(["科目编码", "科目名称", "金额"])
        for item in data.get("assets", {}).get("items", []):
            writer.writerow([item["code"], item["name"], item["balance"]])
        writer.writerow(["", "资产合计", data.get("assets", {}).get("total", 0)])
        writer.writerow([])

        # 负债部分
        writer.writerow(["负债"])
        writer.writerow(["科目编码", "科目名称", "金额"])
        for item in data.get("liabilities", {}).get("items", []):
            writer.writerow([item["code"], item["name"], item["balance"]])
        writer.writerow(["", "负债合计", data.get("liabilities", {}).get("total", 0)])
        writer.writerow([])

        # 所有者权益部分
        writer.writerow(["所有者权益"])
        writer.writerow(["科目编码", "科目名称", "金额"])
        for item in data.get("equity", {}).get("items", []):
            writer.writerow([item["code"], item["name"], item["balance"]])
        writer.writerow(["", "所有者权益合计", data.get("equity", {}).get("total", 0)])
        writer.writerow([])

        # 汇总
        summary = data.get("summary", {})
        writer.writerow(["汇总"])
        writer.writerow(["资产总计", summary.get("total_assets", 0)])
        writer.writerow(["负债及所有者权益总计", summary.get("total_liabilities_equity", 0)])
        writer.writerow(["是否平衡", "是" if summary.get("is_balanced") else "否"])

        return output.getvalue().encode("utf-8-sig")

    def _export_income_statement_csv(self, data: Dict) -> bytes:
        """导出利润表 CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([data.get("report_name", "利润表")])
        writer.writerow([f"报告期间: {data.get('period', '')}"])
        writer.writerow([f"生成时间: {data.get('generated_at', '')}"])
        writer.writerow([])

        # 收入
        writer.writerow(["一、营业收入"])
        writer.writerow(["科目编码", "科目名称", "金额"])
        for item in data.get("revenue", {}).get("items", []):
            writer.writerow([item["code"], item["name"], item["amount"]])
        writer.writerow(["", "收入合计", data.get("revenue", {}).get("total", 0)])
        writer.writerow([])

        # 成本
        writer.writerow(["二、营业成本"])
        writer.writerow(["科目编码", "科目名称", "金额"])
        for item in data.get("cost", {}).get("items", []):
            writer.writerow([item["code"], item["name"], item["amount"]])
        writer.writerow(["", "成本合计", data.get("cost", {}).get("total", 0)])
        writer.writerow([])

        # 费用
        writer.writerow(["三、期间费用"])
        writer.writerow(["科目编码", "科目名称", "金额"])
        for item in data.get("expenses", {}).get("items", []):
            writer.writerow([item["code"], item["name"], item["amount"]])
        writer.writerow(["", "费用合计", data.get("expenses", {}).get("total", 0)])
        writer.writerow([])

        # 汇总
        summary = data.get("summary", {})
        writer.writerow(["利润汇总"])
        writer.writerow(["营业收入", summary.get("total_revenue", 0)])
        writer.writerow(["营业成本", summary.get("total_cost", 0)])
        writer.writerow(["毛利润", summary.get("gross_profit", 0)])
        writer.writerow(["期间费用", summary.get("total_expenses", 0)])
        writer.writerow(["营业利润", summary.get("operating_profit", 0)])
        writer.writerow(["净利润", summary.get("net_profit", 0)])
        writer.writerow(["利润率(%)", summary.get("profit_margin", 0)])

        return output.getvalue().encode("utf-8-sig")

    def _export_cash_flow_csv(self, data: Dict) -> bytes:
        """导出现金流量表 CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([data.get("report_name", "现金流量表")])
        writer.writerow([f"报告期间: {data.get('period', '')}"])
        writer.writerow([f"生成时间: {data.get('generated_at', '')}"])
        writer.writerow([])

        # 经营活动
        operating = data.get("operating_activities", {})
        writer.writerow(["一、经营活动产生的现金流量"])
        writer.writerow(["现金流入", operating.get("inflow", 0)])
        writer.writerow(["现金流出", operating.get("outflow", 0)])
        writer.writerow(["经营活动净现金流", operating.get("net", 0)])
        writer.writerow([])

        # 投资活动
        investing = data.get("investing_activities", {})
        writer.writerow(["二、投资活动产生的现金流量"])
        writer.writerow(["现金流入", investing.get("inflow", 0)])
        writer.writerow(["现金流出", investing.get("outflow", 0)])
        writer.writerow(["投资活动净现金流", investing.get("net", 0)])
        writer.writerow([])

        # 筹资活动
        financing = data.get("financing_activities", {})
        writer.writerow(["三、筹资活动产生的现金流量"])
        writer.writerow(["现金流入", financing.get("inflow", 0)])
        writer.writerow(["现金流出", financing.get("outflow", 0)])
        writer.writerow(["筹资活动净现金流", financing.get("net", 0)])
        writer.writerow([])

        # 汇总
        summary = data.get("summary", {})
        writer.writerow(["现金流量汇总"])
        writer.writerow(["现金净增加额", summary.get("net_cash_flow", 0)])
        writer.writerow(["期初现金余额", summary.get("opening_cash", 0)])
        writer.writerow(["期末现金余额", summary.get("closing_cash", 0)])
        writer.writerow(["是否核对一致", "是" if summary.get("is_reconciled") else "否"])

        return output.getvalue().encode("utf-8-sig")

    def _export_account_balances_csv(self, data: List[Dict]) -> bytes:
        """导出科目余额表 CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["科目余额表"])
        writer.writerow([])
        writer.writerow([
            "科目编码", "科目名称", "类型", "方向", "级次",
            "期初借方", "期初贷方",
            "本期借方", "本期贷方",
            "年累计借方", "年累计贷方",
            "期末借方", "期末贷方"
        ])

        for item in data:
            writer.writerow([
                item.get("code", ""),
                item.get("name", ""),
                item.get("type", ""),
                item.get("direction", ""),
                item.get("level", ""),
                item.get("opening_debit", 0),
                item.get("opening_credit", 0),
                item.get("period_debit", 0),
                item.get("period_credit", 0),
                item.get("ytd_debit", 0),
                item.get("ytd_credit", 0),
                item.get("closing_debit", 0),
                item.get("closing_credit", 0)
            ])

        return output.getvalue().encode("utf-8-sig")

    def _export_ledger_csv(self, data: Dict) -> bytes:
        """导出明细账 CSV"""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([f"明细账 - {data.get('account_code', '')} {data.get('account_name', '')}"])
        writer.writerow([f"期间: {data.get('period', '')}"])
        writer.writerow([])
        writer.writerow(["日期", "凭证类型", "凭证号", "摘要", "借方", "贷方", "余额", "方向"])

        for entry in data.get("entries", []):
            writer.writerow([
                entry.get("date", ""),
                entry.get("voucher_type", ""),
                entry.get("voucher_number", ""),
                entry.get("summary", ""),
                entry.get("debit", 0),
                entry.get("credit", 0),
                entry.get("balance", 0),
                entry.get("direction", "")
            ])

        writer.writerow([])
        summary = data.get("summary", {})
        writer.writerow(["合计", "", "", "", summary.get("total_debit", 0), summary.get("total_credit", 0), summary.get("closing_balance", 0), ""])

        return output.getvalue().encode("utf-8-sig")

    # ==================== Excel 导出 ====================

    def export_to_excel(self, data: Dict[str, Any], report_type: str) -> bytes:
        """
        导出报表为 Excel 格式

        需要安装 openpyxl: pip install openpyxl
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            log.warning("openpyxl not installed, falling back to CSV")
            return self.export_to_csv(data, report_type)

        wb = Workbook()
        ws = wb.active

        # 样式定义
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        if report_type == "balance-sheet":
            self._build_balance_sheet_excel(ws, data, title_font, header_font, header_fill, thin_border)
        elif report_type == "income-statement":
            self._build_income_statement_excel(ws, data, title_font, header_font, header_fill, thin_border)
        elif report_type == "cash-flow":
            self._build_cash_flow_excel(ws, data, title_font, header_font, header_fill, thin_border)
        elif report_type == "account-balances":
            self._build_account_balances_excel(ws, data, title_font, header_font, header_fill, thin_border)
        else:
            raise ValueError(f"不支持的报表类型: {report_type}")

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _build_balance_sheet_excel(self, ws, data, title_font, header_font, header_fill, border):
        """构建资产负债表 Excel"""
        from openpyxl.utils import get_column_letter

        ws.title = "资产负债表"

        # 标题
        ws.merge_cells('A1:D1')
        ws['A1'] = data.get("report_name", "资产负债表")
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"报告期间: {data.get('period', '')}"
        ws['A3'] = f"生成时间: {data.get('generated_at', '')}"

        row = 5

        # 资产部分
        ws[f'A{row}'] = "资产"
        ws[f'A{row}'].font = header_font
        row += 1

        ws[f'A{row}'] = "科目编码"
        ws[f'B{row}'] = "科目名称"
        ws[f'C{row}'] = "金额"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = border
        row += 1

        for item in data.get("assets", {}).get("items", []):
            ws[f'A{row}'] = item["code"]
            ws[f'B{row}'] = item["name"]
            ws[f'C{row}'] = item["balance"]
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1

        ws[f'B{row}'] = "资产合计"
        ws[f'C{row}'] = data.get("assets", {}).get("total", 0)
        ws[f'B{row}'].font = header_font
        ws[f'C{row}'].font = header_font
        row += 2

        # 负债部分
        ws[f'A{row}'] = "负债"
        ws[f'A{row}'].font = header_font
        row += 1

        ws[f'A{row}'] = "科目编码"
        ws[f'B{row}'] = "科目名称"
        ws[f'C{row}'] = "金额"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = border
        row += 1

        for item in data.get("liabilities", {}).get("items", []):
            ws[f'A{row}'] = item["code"]
            ws[f'B{row}'] = item["name"]
            ws[f'C{row}'] = item["balance"]
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1

        ws[f'B{row}'] = "负债合计"
        ws[f'C{row}'] = data.get("liabilities", {}).get("total", 0)
        ws[f'B{row}'].font = header_font
        ws[f'C{row}'].font = header_font
        row += 2

        # 所有者权益部分
        ws[f'A{row}'] = "所有者权益"
        ws[f'A{row}'].font = header_font
        row += 1

        ws[f'A{row}'] = "科目编码"
        ws[f'B{row}'] = "科目名称"
        ws[f'C{row}'] = "金额"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = border
        row += 1

        for item in data.get("equity", {}).get("items", []):
            ws[f'A{row}'] = item["code"]
            ws[f'B{row}'] = item["name"]
            ws[f'C{row}'] = item["balance"]
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1

        ws[f'B{row}'] = "所有者权益合计"
        ws[f'C{row}'] = data.get("equity", {}).get("total", 0)
        ws[f'B{row}'].font = header_font
        ws[f'C{row}'].font = header_font

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18

    def _build_income_statement_excel(self, ws, data, title_font, header_font, header_fill, border):
        """构建利润表 Excel"""
        ws.title = "利润表"

        ws.merge_cells('A1:D1')
        ws['A1'] = data.get("report_name", "利润表")
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"报告期间: {data.get('period', '')}"

        row = 4
        sections = [
            ("一、营业收入", "revenue"),
            ("二、营业成本", "cost"),
            ("三、期间费用", "expenses")
        ]

        for section_name, section_key in sections:
            ws[f'A{row}'] = section_name
            ws[f'A{row}'].font = header_font
            row += 1

            ws[f'A{row}'] = "科目编码"
            ws[f'B{row}'] = "科目名称"
            ws[f'C{row}'] = "金额"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
                ws[f'{col}{row}'].border = border
            row += 1

            for item in data.get(section_key, {}).get("items", []):
                ws[f'A{row}'] = item["code"]
                ws[f'B{row}'] = item["name"]
                ws[f'C{row}'] = item["amount"]
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                row += 1

            ws[f'B{row}'] = "小计"
            ws[f'C{row}'] = data.get(section_key, {}).get("total", 0)
            ws[f'B{row}'].font = header_font
            row += 2

        # 汇总
        summary = data.get("summary", {})
        ws[f'A{row}'] = "利润汇总"
        ws[f'A{row}'].font = header_font
        row += 1
        ws[f'A{row}'] = "净利润"
        ws[f'B{row}'] = summary.get("net_profit", 0)
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18

    def _build_cash_flow_excel(self, ws, data, title_font, header_font, header_fill, border):
        """构建现金流量表 Excel"""
        ws.title = "现金流量表"

        ws.merge_cells('A1:C1')
        ws['A1'] = data.get("report_name", "现金流量表")
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A2'] = f"报告期间: {data.get('period', '')}"

        row = 4
        activities = [
            ("一、经营活动产生的现金流量", "operating_activities"),
            ("二、投资活动产生的现金流量", "investing_activities"),
            ("三、筹资活动产生的现金流量", "financing_activities")
        ]

        for section_name, section_key in activities:
            ws[f'A{row}'] = section_name
            ws[f'A{row}'].font = header_font
            row += 1

            activity = data.get(section_key, {})
            ws[f'A{row}'] = "现金流入"
            ws[f'B{row}'] = activity.get("inflow", 0)
            row += 1
            ws[f'A{row}'] = "现金流出"
            ws[f'B{row}'] = activity.get("outflow", 0)
            row += 1
            ws[f'A{row}'] = "净现金流"
            ws[f'B{row}'] = activity.get("net", 0)
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = header_font
            row += 2

        summary = data.get("summary", {})
        ws[f'A{row}'] = "现金净增加额"
        ws[f'B{row}'] = summary.get("net_cash_flow", 0)
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18

    def _build_account_balances_excel(self, ws, data, title_font, header_font, header_fill, border):
        """构建科目余额表 Excel"""
        ws.title = "科目余额表"

        ws.merge_cells('A1:M1')
        ws['A1'] = "科目余额表"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        headers = [
            "科目编码", "科目名称", "类型", "方向", "级次",
            "期初借方", "期初贷方", "本期借方", "本期贷方",
            "年累计借方", "年累计贷方", "期末借方", "期末贷方"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row += 1
        for item in data:
            values = [
                item.get("code", ""),
                item.get("name", ""),
                item.get("type", ""),
                item.get("direction", ""),
                item.get("level", ""),
                item.get("opening_debit", 0),
                item.get("opening_credit", 0),
                item.get("period_debit", 0),
                item.get("period_credit", 0),
                item.get("ytd_debit", 0),
                item.get("ytd_credit", 0),
                item.get("closing_debit", 0),
                item.get("closing_credit", 0)
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
            row += 1

        # 调整列宽
        widths = [12, 20, 10, 8, 6, 12, 12, 12, 12, 12, 12, 12, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width


# 单例
_export_service = None


def get_export_service() -> ReportExportService:
    """获取导出服务单例"""
    global _export_service
    if _export_service is None:
        _export_service = ReportExportService()
    return _export_service


def get_column_letter(col_idx):
    """获取列字母 (兼容无 openpyxl 情况)"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result
